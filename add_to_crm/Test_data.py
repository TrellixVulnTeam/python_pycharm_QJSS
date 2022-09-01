import _thread
import requests
import time
import datetime
import openpyxl
import pandas as pd

def sheet_method():
    wk = openpyxl.load_workbook('D:\\钉钉文件\\交付记录\\交付记录02.xlsx')
    wk.remove(wk.worksheets[0])
    wk.create_sheet(title='sheet1',index=0)
    wk.close()
    wk.save('D:\\钉钉文件\\交付记录\\交付记录02.xlsx')

# 找企业高筛获取企业id（entity_ids）
url = 'https://sales-test.tangees.com/api/advance-search'
headers = {
    "Cookie":"DISTINCT_ID=a441cf24-6758-4c90-9675-bd0f807de112; __last_enter_version=sales-test; SecurityCenterDuId=IllON0ZMcFlFMjBlRWFycitGbFFHYVdRPSI.FdNl7Q.HXP0Wk5lgP_UJcINFBiozJRvjZ0; _co_i=623943dce572f06e41788438; sessionId=.eJw1zk0KQjEMBOC7dO2iSZqX1stI84cuFHnqSry7BXE1A8PA9y6n3ONxLsfn_opDOV28HIsjTk4H8JikIq2pGE6oVsXU2GdSTPe-jewSJN16n9ZAgddYe0OtREpQhZ3W3VemZfBIi8YbDmLL2RJQuW5gYqMO1I46OMqC3GO_zlvcnn_a6xH7j7cBAYty2hjea67amiCVzxfJkDwq.FdOi9g.sfCB7iEX35wda7-ntLJHbriPS7Q; accountCenterSessionId=.eJw9jt1KAzEQRt8l173I5GeS6XWLCK4FrZTuzTJJJrRuu0J3tRbx3V0UvD2c7-N8qUGkdCN_SDe9dSWpZeXTKAvVHYtaKkEbIwbKHIIkRhJvIrsUY4jGkFB2xJLE2wqSvdSZI2WdIfvsmFwBTzPTlQUponYeIFsL1hmrPbDLWKEGTBEog-eqiQIFl4BRQlUL9T7K5S8GTYXisIgPpmrQhckSgsxSVy8yHv7jx1-_XT0e27unfm_W02bb-_ZZ6_bcwMO2vzW7l2l_bm6b3f7avt7bZrX-nH_ygYdBTvP4Kkl9_wDwxFSC.FdOjMQ.VUAeQh_D2_wd1iwqstlj1a-1cMo"
}
payload = {"filter": "{\"must\":[{\"hasPhone\":[{\"exist\":\"1\"}]}]}", "relation": "0", "start": "460", "end": "480",
           "filter_advance_search_filter_pool": "1", "mode": "2"}
res = requests.post(url, data=payload, headers=headers)
# print(res.text)
# print(res.json())
res_json = res.json()['data']
num = res.json()['size']
print(num)
en_list = []
enname_list = []
for i in range(num):
    en_list.append(res_json[i]['_id'])
    enname_list.append(res_json[i]['name'])
#print(en_list)
#print(enname_list)
#删除原有表格与数据
sheet_method()
filename = "交付记录02.xlsx"
xfile = openpyxl.load_workbook('D:\\钉钉文件\\交付记录\\%s'%filename)  # 加载文件
sheet1 = xfile.worksheets[0]
#xfile.remove(sheet1)
sheet1.cell(1, 1).value = "entity_id"
sheet1.cell(1, 2).value = "entity_name"
sheet1.cell(1, 3).value = "entity_type"
sheet1.cell(1, 4).value = "rate"
for i in range(len(en_list)):
    sheet1.cell(i + 2, 1).value = en_list[i]
    sheet1.cell(i + 2, 2).value = enname_list[i]
    sheet1.cell(i + 2, 3).value = "enterprise"
    sheet1.cell(i + 2, 4).value = "1"
# 保存数据，如果提示权限错误，需要关闭打开的excel
xfile.save('D:\\钉钉文件\\交付记录\\%s'%filename)   #保存文件

#xlsx文件格式转换为csv格式
data_xls = pd.read_excel('D:\\钉钉文件\\交付记录\\%s'%filename, index_col=0)
data_xls.to_csv('D:\\钉钉文件\\交付记录\\交付记录02.csv', encoding='utf-8')
