#!/usr/bin/env python
#coding:utf-8

import requests
import random
import openpyxl
import os
import sys
'''
r=requests.get(
    url='http://yuedu.baidu.com/ebook/3c0077aaa32d7375a41780bb',
    params={'_searchquery':'selenium-python%D7%D4%B6%AF%BB%AF%B2%E2%CA%D4'})
print (r.url)
'''


'''
r=requests.post(
    url='http://m.cyw.com/index.php?m=api&c=cookie&a=setcity',
    data={'cityId':438})
print (r.json())

r=requests.get('http://www.bing.com')
print (u'HTTP状态码:',r.status_code)
print (u'请求的URL:',r.url)
print (u'获取Headers:',r.headers)
print (u'响应内容:',r.text)

# -*- coding:utf-8 -*-
def Create_num():
  list = ['139','138','137','136','135','134','159','158','157','150','151','152','188','187','182','183','184','178','130','131','132','156','155','186','185','176','133','153','189','180','181','177']
  str = '0123456789'
  for i in range(100):
    print(random.choice(list) + "".join(random.choice(str) for i in range(8)))
Create_num()



def create_a_phone():
    # 第二位数字
    second = [3, 4, 5, 7, 8][random.randint(0, 4)]

    # 第三位数字
    third = {3: random.randint(0, 9),
             4: [5, 7, 9][random.randint(0, 2)],
             5: [i for i in range(10) if i != 4][random.randint(0, 8)],
             7: [i for i in range(10) if i not in [4, 9]][random.randint(0, 7)],
             8: random.randint(0, 9), }[second]

    # 最后八位数字
    suffix = random.randint(9999999, 100000000)

    # 拼接手机号
    print("1{}{}{}".format(second, third, suffix))
for i in range(100):
 create_a_phone()

#随机生成后八位数（生成号码）
def creat():
    eight = ''
    for i in range(8):
        eight=eight+str(random.randint(0,9))
    return "135{}".format(eight)

#随机生成号码（前三位有多种情况）
def Create_num():
  list = ['139','138','137','136','135','134','159','158','157','150','151','152','188','187','182','183','184','178','130','131','132','156','155','186','185','176','133','153','189','180','181','177']
  str = '0123456789'
  return (random.choice(list) + "".join(random.choice(str) for i in range(8)))
'''
'''
#生成号码(改进)且导入到智能呼叫号码池中
def create_a_phone():
    # 第二位数字
    second = [3, 4, 5, 7, 8][random.randint(0, 4)]

    # 第三位数字
    third = {3: random.randint(0, 9),
             4: [5, 7, 9][random.randint(0, 2)],
             5: [i for i in range(10) if i != 4][random.randint(0, 8)],
             7: [i for i in range(10) if i not in [4, 9]][random.randint(0, 7)],
             8: random.randint(0, 9), }[second]

    # 最后八位数字
    suffix = random.randint(9999999, 100000000)

    # 拼接手机号
    return ("1{}{}{}".format(second, third, suffix))
name='0'
user='测试浮窗'
list=[] #存放每条手机号和密码
user_list=[] #存放所有的手机号和密码
for i in range(200):
    list=[user,name,create_a_phone()]
    user_list.append(list)
    name =int (name)+1

book=openpyxl.Workbook()
sheet1=book.active
a=['联系人','企业名称','被叫号码']
sheet1.append(a)
for i in range(len(user_list)):
    sheet1.append(user_list[i])
book.save('D:/钉钉文件/号码包/改进号码包200(1).xlsx')
'''
'''
#生成用户导入到部门与用户中
def create_a_phone():
    # 第二位数字
    second = [3, 4, 5, 7, 8][random.randint(0, 4)]

    # 第三位数字
    third = {3: random.randint(0, 9),
             4: [5, 7, 9][random.randint(0, 2)],
             5: [i for i in range(10) if i != 4][random.randint(0, 8)],
             7: [i for i in range(10) if i not in [4, 9]][random.randint(0, 7)],
             8: random.randint(0, 9), }[second]

    # 最后八位数字
    suffix = random.randint(9999999, 100000000)

    # 拼接手机号
    return ("1{}{}{}".format(second, third, suffix))
name='HH'
user='超级管理员'
email=''
department='测试'
ZG=''
bz=''
'''
'''
A='T'
for i in range(10):
   T = A + str(i)

   print(T)
'''
'''
T1='否'
T2='否'
T3='否'
T4='否'
T5='否'
T6='否'
T7='是'
T8='否'
T9='否'
T10='否'
list=[] #存放每条手机号和密码
user_list=[] #存放所有的手机号和密码
for i in range(10):
    list=[create_a_phone(),name,email,user,department,ZG,bz,T1,T2,T3,T4,T5,T6,T7,T8,T9,T10]
    user_list.append(list)

book=openpyxl.Workbook()
sheet1=book.active
a=['手机号码','姓名','邮箱','所属角色','所属部门','主管','备注','企业中心是否开通','探迹拓客是否开通','探迹触达是否开通','探迹CRM是否开通','探迹智能呼叫是否开通','探迹拓客VIP是否开通','探迹开放平台是否开通','探迹集客是否开通','探迹企微SCRM是否开通','探迹营销是否开通']
sheet1.append(a)
for i in range(len(user_list)):
    sheet1.append(user_list[i])
book.save('D:/钉钉文件/部门与角色导入用户/导入用户10个.xlsx')
'''
'''
#重新创建一个excel文件导入客户
def create_a_phone():
    # 第二位数字
    second = [3, 4, 5, 7, 8][random.randint(0, 4)]

    # 第三位数字
    third = {3: random.randint(0, 9),
             4: [5, 7, 9][random.randint(0, 2)],
             5: [i for i in range(10) if i != 4][random.randint(0, 8)],
             7: [i for i in range(10) if i not in [4, 9]][random.randint(0, 7)],
             8: random.randint(0, 9), }[second]

    # 最后八位数字
    suffix = random.randint(9999999, 100000000)

    # 拼接手机号
    return ("1{}{}{}".format(second, third, suffix))
name=random.randint(0,999999)
user='广州探迹科技有限公司'
status='有意向'
T1=''
T2=''
T3=''
T4=''
T5=''
T6=''
T7=''
T8=''
T9=''
list=[] #存放每条手机号和密码
user_list=[] #存放所有的手机号和密码
for i in range(19000):
    username = '测试' + str(name)
    phone = random.randint(1000000, 10000000)
    list=[username,user,T1,phone,create_a_phone(),T2,T3,T4,T5,T6,T7,T8,T9,status]
    user_list.append(list)
    name =int (name)+1

book=openpyxl.Workbook()
sheet1=book.active
a=['备注：表头中标红字段为必填字段']
b=['客户名称','公司名称','客户类型','电话','手机','微信号','QQ号','旺旺号','邮箱','网址','地区','地址','邮编','跟进状态','客户来源','业务标签','公司规模','备注1111','姓名','注册资本','成立日期','是否高新企业','是否上市','是否有证书','注册地址','所属行业','企业类型','年营业额','人员规模','jun','jun2','jun22'
]
sheet1.append(a)
sheet1.append(b)

for i in range(len(user_list)):
    sheet1.append(user_list[i])
book.save('D:/钉钉文件/导入客户/客户包19000(1).xlsx')
'''
#在已有的excel文件中写入新数据导入客户
def create_a_phone():
    # 第二位数字
    second = [3, 4, 5, 7, 8][random.randint(0, 4)]

    # 第三位数字
    third = {3: random.randint(0, 9),
             4: [5, 7, 9][random.randint(0, 2)],
             5: [i for i in range(10) if i != 4][random.randint(0, 8)],
             7: [i for i in range(10) if i not in [4, 9]][random.randint(0, 7)],
             8: random.randint(0, 9), }[second]

    # 最后八位数字
    suffix = random.randint(9999999, 100000000)

    # 拼接手机号
    return ("1{}{}{}".format(second, third, suffix))
list=''
user_list=[]
user_phone=[]
Name=[]
number=random.randint(0,999999)

for i in range(200):
    name='广州探迹'+str(number)
    phone = random.randint(1000000, 10000000)
    number = int(number) + 1
    list=create_a_phone()
    user_phone.append(phone)
    user_list.append(list)
    Name.append(name)
N=Name
P=user_phone
L=user_list
user_name=''
status='成交'
origin='其他'
##写入xlsx文件
xfile = openpyxl.load_workbook('D:\\钉钉文件\\导入客户\\客户包5.xlsx')#加载文件
sheet1 = xfile.worksheets[0]

for i in range(len(L)):
    sheet1.cell(i+3,1).value=N[i]
    sheet1.cell(i + 3, 2).value = user_name
    sheet1.cell(i+3, 4).value = P[i]
    sheet1.cell(i+3, 5).value=L[i]
    sheet1.cell(i+3, 14).value = status
    sheet1.cell(i+3,15).value = origin
# 保存数据，如果提示权限错误，需要关闭打开的excel
xfile.save('D:\\钉钉文件\\导入客户\\客户包5.xlsx')
